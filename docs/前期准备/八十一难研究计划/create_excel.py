import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

def create_excel_analysis():
    # 读取CSV文件
    df = pd.read_csv('西游记八十一难对照表.csv')
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    if wb.active:
        wb.remove(wb.active)
    
    # 创建主要数据表
    ws1 = wb.create_sheet('八十一难详表')
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 应用标题样式
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 设置列宽
    column_widths = [8, 20, 25, 20, 50, 20, 25]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # 为所有数据单元格添加边框
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 创建人物统计表
    ws2 = wb.create_sheet('主要人物统计')
    
    # 统计主要人物出现频次
    all_characters = []
    for characters in df['主要人物']:
        if pd.notna(characters):
            chars = [char.strip() for char in characters.split('、')]
            all_characters.extend(chars)
    
    char_counts = pd.Series(all_characters).value_counts()
    
    # 写入人物统计数据
    ws2.append(['人物名称', '出现次数', '占比'])
    total_appearances = len(all_characters)
    
    for char, count in char_counts.head(20).items():  # 取前20个
        percentage = f"{count/total_appearances*100:.1f}%"
        ws2.append([char, count, percentage])
    
    # 设置人物统计表样式
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    
    # 创建地点统计表
    ws3 = wb.create_sheet('地点统计')
    
    # 统计地点出现频次
    all_locations = []
    for locations in df['地点']:
        if pd.notna(locations):
            locs = [loc.strip() for loc in locations.split('、')]
            all_locations.extend(locs)
    
    loc_counts = pd.Series(all_locations).value_counts()
    
    # 写入地点统计数据
    ws3.append(['地点名称', '出现次数', '占比'])
    total_locations = len(all_locations)
    
    for loc, count in loc_counts.head(15).items():  # 取前15个
        percentage = f"{count/total_locations*100:.1f}%"
        ws3.append([loc, count, percentage])
    
    # 设置地点统计表样式
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    
    # 创建文化内涵分析表
    ws4 = wb.create_sheet('文化内涵分析')
    
    # 统计文化内涵关键词
    cultural_keywords = []
    for content in df['文化内涵']:
        if pd.notna(content):
            # 提取关键词
            keywords = ['佛教', '道教', '儒家', '因果', '慈悲', '智慧', '降魔', '修行', '团队', '正义']
            for keyword in keywords:
                if keyword in content:
                    cultural_keywords.append(keyword)
    
    cultural_counts = pd.Series(cultural_keywords).value_counts()
    
    # 写入文化内涵分析数据
    ws4.append(['文化主题', '出现次数', '占比', '代表性难次'])
    
    for theme, count in cultural_counts.items():
        percentage = f"{count/len(cultural_keywords)*100:.1f}%"
        # 找到代表性的难次
        representative_trials = []
        for idx, content in enumerate(df['文化内涵']):
            if pd.notna(content) and theme in content:
                representative_trials.append(str(df.iloc[idx]['难次']))
                if len(representative_trials) >= 3:  # 只取前3个
                    break
        rep_trials_str = '、'.join(representative_trials)
        ws4.append([theme, count, percentage, rep_trials_str])
    
    # 设置文化内涵分析表样式
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 20
    
    # 创建象征意义分析表
    ws5 = wb.create_sheet('象征意义分析')
    
    # 统计象征意义关键词
    symbolic_keywords = []
    for content in df['象征意义']:
        if pd.notna(content):
            # 提取关键词
            keywords = ['考验', '成长', '智慧', '团队', '修行', '磨难', '感化', '正义', '慈悲', '降魔']
            for keyword in keywords:
                if keyword in content:
                    symbolic_keywords.append(keyword)
    
    symbolic_counts = pd.Series(symbolic_keywords).value_counts()
    
    # 写入象征意义分析数据
    ws5.append(['象征主题', '出现次数', '占比', '代表性难次'])
    
    for theme, count in symbolic_counts.items():
        percentage = f"{count/len(symbolic_keywords)*100:.1f}%"
        # 找到代表性的难次
        representative_trials = []
        for idx, content in enumerate(df['象征意义']):
            if pd.notna(content) and theme in content:
                representative_trials.append(str(df.iloc[idx]['难次']))
                if len(representative_trials) >= 3:  # 只取前3个
                    break
        rep_trials_str = '、'.join(representative_trials)
        ws5.append([theme, count, percentage, rep_trials_str])
    
    # 设置象征意义分析表样式
    for cell in ws5[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws5.column_dimensions['A'].width = 15
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 20
    
    # 创建研究摘要表
    ws6 = wb.create_sheet('研究摘要')
    
    # 添加研究摘要内容
    summary_data = [
        ['研究项目', '《西游记》九九八十一难系统研究'],
        ['数据来源', '《西游记》原著文本及相关研究文献'],
        ['研究范围', '八十一难的完整梳理与分析'],
        ['', ''],
        ['统计概览', ''],
        ['总难次数', '81'],
        ['涉及主要人物', f'{len(char_counts)}个'],
        ['涉及地点', f'{len(loc_counts)}个'],
        ['文化主题', f'{len(cultural_counts)}个'],
        ['象征意义', f'{len(symbolic_counts)}个'],
        ['', ''],
        ['核心发现', ''],
        ['最频繁出现人物', f'{char_counts.index[0]}（{char_counts.iloc[0]}次）'],
        ['最常见地点', f'{loc_counts.index[0]}（{loc_counts.iloc[0]}次）'],
        ['主要文化主题', f'{cultural_counts.index[0]}（{cultural_counts.iloc[0]}次）'],
        ['核心象征意义', f'{symbolic_counts.index[0]}（{symbolic_counts.iloc[0]}次）'],
        ['', ''],
        ['研究价值', ''],
        ['学术价值', '为《西游记》研究提供系统化数据支撑'],
        ['文化价值', '展现中国传统文化的多元融合特征'],
        ['教育价值', '为文学教育和文化传承提供参考'],
        ['', ''],
        ['建议进一步研究', ''],
        ['深度分析', '每一难的详细文本分析'],
        ['比较研究', '与其他版本《西游记》的对比'],
        ['跨文化研究', '八十一难在不同文化中的传播与接受'],
        ['现代应用', '八十一难在现代文艺作品中的改编与应用']
    ]
    
    for row_data in summary_data:
        ws6.append(row_data)
    
    # 设置研究摘要表样式
    for row in ws6.iter_rows():
        for cell in row:
            if cell.value and ('研究项目' in str(cell.value) or '统计概览' in str(cell.value) or
                              '核心发现' in str(cell.value) or '研究价值' in str(cell.value) or
                              '建议进一步研究' in str(cell.value)):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 50
    
    # 保存Excel文件
    wb.save('西游记八十一难对照表.xlsx')
    print("Excel文件创建成功！")
    
    # 输出统计信息
    print(f"\n=== 数据统计摘要 ===")
    print(f"总难次数: 81")
    print(f"涉及主要人物: {len(char_counts)}个")
    print(f"涉及地点: {len(loc_counts)}个")
    print(f"文化主题: {len(cultural_counts)}个")
    print(f"象征意义: {len(symbolic_counts)}个")
    print(f"\n最频繁出现的人物: {char_counts.head(5).to_dict()}")
    print(f"\n最常见的地点: {loc_counts.head(5).to_dict()}")
    print(f"\n主要文化主题: {cultural_counts.to_dict()}")
    print(f"\n核心象征意义: {symbolic_counts.to_dict()}")

if __name__ == "__main__":
    create_excel_analysis()